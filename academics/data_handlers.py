import csv
import io
from decimal import Decimal, InvalidOperation
from datetime import datetime

from django.contrib.auth import get_user_model
from django.utils.dateparse import parse_date

from .models import (
    StudentProfile, TeacherProfile, ParentProfile,
    Class, Subject, Exam, Result,
)

User = get_user_model()


# ─────────────────────────────────────────────
# EXPORT HELPERS
# ─────────────────────────────────────────────

def _student_rows(school=None):
    qs = StudentProfile.objects.select_related('student_class', 'user').all()
    if school:
        qs = qs.filter(school=school)
    rows = []
    for s in qs:
        rows.append({
            'full_name': s.full_name,
            'guardian_name': s.guardian_name,
            'date_of_birth': s.date_of_birth.strftime('%Y-%m-%d') if s.date_of_birth else '',
            'gender': s.get_gender_display(),
            'roll_number': s.roll_number or '',
            'admission_number': s.admission_number or '',
            'class_name': s.student_class.name if s.student_class else '',
            'class_section': s.student_class.section if s.student_class else '',
            'phone': s.phone,
            'address': s.address,
            'email': s.email or (s.user.email if s.user else ''),
        })
    return rows


def _teacher_rows(school=None):
    qs = TeacherProfile.objects.select_related('user').all()
    if school:
        qs = qs.filter(school=school)
    rows = []
    for t in qs:
        rows.append({
            'full_name': t.full_name,
            'father_name': t.father_name,
            'date_of_birth': t.date_of_birth.strftime('%Y-%m-%d') if t.date_of_birth else '',
            'gender': t.get_gender_display(),
            'employee_id': t.employee_id or '',
            'cnic': t.cnic,
            'designation': t.get_designation_display(),
            'employment_type': t.get_employment_type_display(),
            'phone': t.phone,
            'email': t.email or (t.user.email if t.user else ''),
            'salary': str(t.salary) if t.salary else '',
            'bank_name': t.bank_name,
            'bank_account': t.bank_account,
        })
    return rows


def _parent_rows(school=None):
    qs = ParentProfile.objects.select_related('user').prefetch_related('students').all()
    if school:
        qs = qs.filter(school=school)
    rows = []
    for p in qs:
        student_nums = ', '.join(
            s.admission_number for s in p.students.all() if s.admission_number
        )
        rows.append({
            'full_name': p.full_name,
            'relation': p.get_relation_display(),
            'phone': p.phone,
            'email': p.email or (p.user.email if p.user else ''),
            'student_admission_numbers': student_nums,
        })
    return rows


def _result_rows(school=None):
    qs = Result.objects.select_related(
        'student', 'exam', 'exam__subject', 'exam__assigned_class'
    ).all()
    if school:
        qs = qs.filter(student__school=school)
    rows = []
    for r in qs:
        rows.append({
            'student_name': r.student.full_name,
            'admission_number': r.student.admission_number or '',
            'exam_name': r.exam.name,
            'subject_code': r.exam.subject.code,
            'class_name': r.exam.assigned_class.name,
            'class_section': r.exam.assigned_class.section,
            'total_marks': r.exam.total_marks,
            'passing_marks': r.exam.passing_marks,
            'marks_obtained': r.marks_obtained,
        })
    return rows


STUDENT_HEADERS = ['full_name', 'guardian_name', 'date_of_birth', 'gender', 'roll_number',
                   'admission_number', 'class_name', 'class_section', 'phone', 'address', 'email']
TEACHER_HEADERS = ['full_name', 'father_name', 'date_of_birth', 'gender', 'employee_id',
                   'cnic', 'designation', 'employment_type', 'phone', 'email', 'salary',
                   'bank_name', 'bank_account']
PARENT_HEADERS = ['full_name', 'relation', 'phone', 'email', 'student_admission_numbers']
RESULT_HEADERS = ['student_name', 'admission_number', 'exam_name', 'subject_code',
                  'class_name', 'class_section', 'total_marks', 'passing_marks', 'marks_obtained']

EXPORTERS = {
    'students':  {'headers': STUDENT_HEADERS,  'getter': _student_rows},
    'teachers':  {'headers': TEACHER_HEADERS,  'getter': _teacher_rows},
    'parents':   {'headers': PARENT_HEADERS,   'getter': _parent_rows},
    'results':   {'headers': RESULT_HEADERS,   'getter': _result_rows},
}


def export_to_csv(data_type, school=None):
    info = EXPORTERS[data_type]
    rows = info['getter'](school)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=info['headers'])
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue()


def export_to_excel(data_type, school=None):
    from openpyxl import Workbook
    info = EXPORTERS[data_type]
    rows = info['getter'](school)

    wb = Workbook()
    ws = wb.active
    ws.title = data_type.capitalize()
    ws.append(info['headers'])
    for row in rows:
        ws.append([row[h] for h in info['headers']])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# IMPORT HELPERS
# ─────────────────────────────────────────────

def _safe_date(val):
    if not val:
        return None
    d = parse_date(str(val).strip())
    return d


def _safe_decimal(val):
    if not val:
        return Decimal('0')
    try:
        return Decimal(str(val).strip())
    except (InvalidOperation, ValueError):
        return Decimal('0')


def _read_file_rows(file_obj):
    """Read CSV or Excel file and return list of dicts (each dict = one row)."""
    filename = getattr(file_obj, 'name', '')
    if filename.lower().endswith(('.xlsx', '.xls')):
        from openpyxl import load_workbook
        wb = load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []
        headers = [str(h).strip() if h else '' for h in rows[0]]
        result = []
        for row in rows[1:]:
            if all(c is None for c in row):
                continue
            row_dict = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                val = row[i] if i < len(row) else None
                row_dict[h] = str(val).strip() if val is not None else ''
            result.append(row_dict)
        return result
    else:
        decoded = file_obj.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(decoded))
        return [row for row in reader]


def _get_or_create_school_user(username, email, role, school=None):
    user = User.objects.filter(username=username).first()
    if user:
        return user, False
    password = f"{username}123"
    user = User.objects.create_user(username=username, email=email, password=password)
    user.role = role
    if school:
        user.school = school
    user.save()
    return user, True


def _resolve_class(name, section, school=None):
    qs = Class.objects.filter(name__iexact=name.strip(), section__iexact=section.strip())
    if school:
        qs = qs.filter(school=school)
    return qs.first()


def _resolve_subject(code):
    return Subject.objects.filter(code__iexact=code.strip()).first()


def _resolve_exam(name, subject_code, class_name, class_section):
    subject = _resolve_subject(subject_code)
    cls = _resolve_class(class_name, class_section)
    if not subject or not cls:
        return None
    return Exam.objects.filter(
        name__iexact=name.strip(), subject=subject, assigned_class=cls
    ).first()


def import_students(file_obj, school=None):
    rows = _read_file_rows(file_obj)
    success = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        try:
            full_name = row.get('full_name', '').strip()
            if not full_name:
                errors.append(f"Row {i}: full_name is required")
                continue

            admission_number = row.get('admission_number', '').strip()
            username = admission_number or full_name.replace(' ', '').lower()
            email = row.get('email', '').strip()

            user, created = _get_or_create_school_user(username, email, 'student', school)

            cls = _resolve_class(
                row.get('class_name', ''), row.get('class_section', ''), school
            )

            StudentProfile.objects.update_or_create(
                admission_number=admission_number,
                defaults={
                    'user': user,
                    'school': school,
                    'full_name': full_name,
                    'guardian_name': row.get('guardian_name', '').strip(),
                    'date_of_birth': _safe_date(row.get('date_of_birth')),
                    'gender': 'M' if 'male' in row.get('gender', 'M').lower() else 'F',
                    'roll_number': row.get('roll_number', '').strip() or None,
                    'student_class': cls,
                    'phone': row.get('phone', '').strip(),
                    'address': row.get('address', '').strip(),
                    'email': email,
                }
            )
            success += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    return {'success': success, 'errors': errors}


def import_teachers(file_obj, school=None):
    rows = _read_file_rows(file_obj)
    success = 0
    errors = []

    DESIGNATION_MAP = {'teacher': 'teacher', 'coordinator': 'coordinator',
                       'manager': 'manager', 'vp': 'vp', 'group head': 'group_head'}
    EMPLOYMENT_MAP = {'permanent': 'permanent', 'contract': 'contract', 'daily wager': 'daily_wager'}

    for i, row in enumerate(rows, start=2):
        try:
            full_name = row.get('full_name', '').strip()
            if not full_name:
                errors.append(f"Row {i}: full_name is required")
                continue

            employee_id = row.get('employee_id', '').strip()
            username = employee_id or full_name.replace(' ', '').lower()
            email = row.get('email', '').strip()

            user, created = _get_or_create_school_user(username, email, 'teacher', school)

            designation_raw = row.get('designation', 'teacher').strip().lower()
            designation = DESIGNATION_MAP.get(designation_raw, 'teacher')
            emp_type_raw = row.get('employment_type', 'permanent').strip().lower()
            employment_type = EMPLOYMENT_MAP.get(emp_type_raw, 'permanent')

            TeacherProfile.objects.update_or_create(
                employee_id=employee_id,
                defaults={
                    'user': user,
                    'school': school,
                    'full_name': full_name,
                    'father_name': row.get('father_name', '').strip(),
                    'date_of_birth': _safe_date(row.get('date_of_birth') or row.get('dob')),
                    'gender': 'M' if 'male' in row.get('gender', 'M').lower() else 'F',
                    'cnic': row.get('cnic', '').strip(),
                    'designation': designation,
                    'employment_type': employment_type,
                    'phone': row.get('phone', '').strip(),
                    'email': email,
                    'joining_date': _safe_date(row.get('joining_date') or row.get('hire_date')),
                    'address': row.get('permanent_address') or row.get('address') or '',
                    'marital_status': row.get('marital_status', '').strip(),
                    'nationality': row.get('nationality', '').strip(),
                    'qualification': row.get('qualification', '').strip(),
                    'blood_group': row.get('blood_group', '').strip(),
                    'religion': row.get('religion', '').strip(),
                    'father_cnic': row.get('father_cnic', '').strip(),
                    'father_occupation': row.get('father_occupation') or row.get('Father-Occupation') or '',
                    'husband_name': row.get('husband_name') or row.get('Husband_name') or '',
                    'husband_cnic': row.get('husband_cnic') or row.get('Husband-CNIC') or '',
                    'husband_occupation': row.get('husband_occupation') or row.get('Husband-Occupation') or '',
                    'salary': _safe_decimal(row.get('salary', '0')),
                    'bank_name': row.get('bank_name', '').strip(),
                    'bank_account': row.get('bank_account', '').strip(),
                }
            )
            success += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    return {'success': success, 'errors': errors}


def import_parents(file_obj, school=None):
    rows = _read_file_rows(file_obj)
    success = 0
    errors = []

    RELATION_MAP = {'father': 'father', 'mother': 'mother', 'guardian': 'guardian'}

    for i, row in enumerate(rows, start=2):
        try:
            full_name = row.get('full_name', '').strip()
            if not full_name:
                errors.append(f"Row {i}: full_name is required")
                continue

            email = row.get('email', '').strip()
            username = full_name.replace(' ', '').lower()
            base = username
            counter = 1
            while User.objects.filter(username=username).exists():
                username = f"{base}{counter}"
                counter += 1

            user, _ = _get_or_create_school_user(username, email, 'parent', school)

            relation_raw = row.get('relation', 'father').strip().lower()
            relation = RELATION_MAP.get(relation_raw, 'father')

            parent, _ = ParentProfile.objects.update_or_create(
                full_name=full_name,
                school=school,
                defaults={
                    'user': user,
                    'relation': relation,
                    'phone': row.get('phone', '').strip(),
                    'email': email,
                }
            )

            nums_str = row.get('student_admission_numbers', '')
            if nums_str:
                nums = [n.strip() for n in nums_str.split(',') if n.strip()]
                students = StudentProfile.objects.filter(admission_number__in=nums)
                parent.students.set(students)

            success += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    return {'success': success, 'errors': errors}


def import_results(file_obj, school=None):
    rows = _read_file_rows(file_obj)
    success = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        try:
            admission_number = row.get('admission_number', '').strip()
            if not admission_number:
                errors.append(f"Row {i}: admission_number is required")
                continue

            student = StudentProfile.objects.filter(admission_number=admission_number).first()
            if not student:
                errors.append(f"Row {i}: Student '{admission_number}' not found")
                continue

            exam = _resolve_exam(
                row.get('exam_name', ''),
                row.get('subject_code', ''),
                row.get('class_name', ''),
                row.get('class_section', ''),
            )
            if not exam:
                errors.append(f"Row {i}: Exam '{row.get('exam_name', '')}' not found for given subject/class")
                continue

            marks = _safe_decimal(row.get('marks_obtained', '0'))

            Result.objects.update_or_create(
                student=student,
                exam=exam,
                defaults={
                    'marks_obtained': float(marks),
                }
            )
            success += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    return {'success': success, 'errors': errors}


IMPORTERS = {
    'students': import_students,
    'teachers': import_teachers,
    'parents':  import_parents,
    'results':  import_results,
}
