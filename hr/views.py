import calendar
import csv
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.utils import timezone
from django.db.models import Q, Sum
import json

from academics.models import TeacherProfile
from .models import EmployeeSalary, MonthlySalary, SalaryConfig, EmployeeAttendance
from .forms import (
    EmployeeSalaryForm, MonthlySalaryForm, SalaryConfigForm,
    EmployeeAttendanceForm, GenerateSalaryForm
)


def get_user_school(user):
    return getattr(user, 'school', None)


# ─────────────────────────────────────────────
# EMPLOYEE (TEACHER) HR VIEW — with all FORMAT.xlsx fields
# ─────────────────────────────────────────────

@login_required(login_url='admin_login')
def employee_list(request):
    """Employee list with search, filter, and CRUD actions."""
    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    school = get_user_school(request.user)
    employees = TeacherProfile.objects.select_related('salary_detail').all()
    if school and not request.user.is_superuser:
        employees = employees.filter(school=school)

    # Search
    search = request.GET.get('search', '').strip()
    if search:
        employees = employees.filter(
            Q(full_name__icontains=search) |
            Q(employee_id__icontains=search) |
            Q(cnic__icontains=search) |
            Q(phone__icontains=search)
        )

    # Filter by employment type
    emp_type = request.GET.get('type', '')
    if emp_type:
        employees = employees.filter(employment_type=emp_type)

    # Filter by designation
    designation = request.GET.get('designation', '')
    if designation:
        employees = employees.filter(designation=designation)

    # Filter by gender
    gender = request.GET.get('gender', '')
    if gender:
        employees = employees.filter(gender=gender)

    # Stats
    total = employees.count()
    permanent = employees.filter(employment_type='permanent').count()
    contract = employees.filter(employment_type='contract').count()
    daily_wager = employees.filter(employment_type='daily_wager').count()
    male = employees.filter(gender='M').count()
    female = employees.filter(gender='F').count()

    # Attach salary info
    for emp in employees:
        sal = getattr(emp, 'salary_detail', None)
        emp.basic_salary = sal.basic_salary if sal else 0
        emp.salary_type_display = sal.get_salary_type_display() if sal else '-'
        emp.employment_type_display = sal.get_employment_type_display() if sal else '-'
        emp.working_days = f"{sal.working_days_per_week}/week" if sal else '-'

    return redirect('admin_console')


# ─────────────────────────────────────────────
# EMPLOYEE DETAIL VIEW
# ─────────────────────────────────────────────

@login_required(login_url='admin_login')
def employee_detail(request, employee_id):
    """View full employee profile with salary history."""
    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    employee = get_object_or_404(TeacherProfile, pk=employee_id)
    emp_salary = EmployeeSalary.objects.filter(employee=employee).first()
    salary_history = MonthlySalary.objects.filter(employee=employee).order_by('-year', '-month')[:12]

    context = {
        'employee': employee,
        'emp_salary': emp_salary,
        'salary_history': salary_history,
        'section': 'employees',
    }
    return render(request, 'hr/employee_detail.html', context)


# ─────────────────────────────────────────────
# EMPLOYEE EDIT
# ─────────────────────────────────────────────

@login_required(login_url='admin_login')
def employee_edit(request, employee_id):
    """Edit employee basic info."""
    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    employee = get_object_or_404(TeacherProfile, pk=employee_id)

    if request.method == 'POST':
        employee.full_name = request.POST.get('full_name', employee.full_name)
        employee.father_name = request.POST.get('father_name', employee.father_name)
        employee.phone = request.POST.get('phone', employee.phone)
        employee.email = request.POST.get('email', employee.email)
        employee.cnic = request.POST.get('cnic', employee.cnic)
        employee.address = request.POST.get('address', employee.address)
        employee.designation = request.POST.get('designation', employee.designation)
        employee.employment_type = request.POST.get('employment_type', employee.employment_type)
        employee.gender = request.POST.get('gender', employee.gender)
        salary_val = request.POST.get('salary', '')
        if salary_val:
            employee.salary = float(salary_val)
        employee.save()

        # Also sync EmployeeSalary basic_salary if it exists
        emp_salary = EmployeeSalary.objects.filter(employee=employee).first()
        if emp_salary and employee.salary > 0:
            emp_salary.basic_salary = employee.salary
            emp_salary.save()

        messages.success(request, f'Employee {employee.full_name} updated successfully.')
        return redirect('admin_console')

    context = {
        'employee': employee,
        'section': 'employees',
    }
    return render(request, 'hr/employee_edit.html', context)


# ─────────────────────────────────────────────
# EMPLOYEE DELETE (Soft Delete)
# ─────────────────────────────────────────────

@login_required(login_url='admin_login')
def employee_delete(request, employee_id):
    """Soft-delete employee (mark as separated)."""
    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    employee = get_object_or_404(TeacherProfile, pk=employee_id)

    if request.method == 'POST':
        employee.is_employee_separated = True
        employee.save()
        messages.success(request, f'{employee.full_name} has been marked as separated.')
        return redirect('admin_console')

    return render(request, 'hr/employee_delete_confirm.html', {
        'employee': employee,
        'section': 'employees',
    })

@login_required(login_url='admin_login')
def salary_config(request):
    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    import calendar as cal
    month = int(request.POST.get('month', timezone.now().month)) if request.method == 'POST' else int(request.GET.get('month', timezone.now().month))
    year = int(request.POST.get('year', timezone.now().year)) if request.method == 'POST' else int(request.GET.get('year', timezone.now().year))

    config = SalaryConfig.objects.filter(month=month, year=year).first()
    if not config:
        config = SalaryConfig.objects.create(month=month, year=year, tax_percentage=0)

    if request.method == 'POST':
        action = request.POST.get('action', 'save')
        if action == 'reset':
            config.delete()
            SalaryConfig.objects.create(month=month, year=year)
            messages.success(request, 'Salary configuration reset to defaults.')
            return redirect(f'/admin-console/?section=salary-config&month={month}&year={year}')

        config.month = month
        config.year = year
        config.default_working_days = int(request.POST.get('default_working_days', 26))
        config.max_allowed_leaves = int(request.POST.get('max_allowed_leaves', 0))
        config.late_deduction_per = int(request.POST.get('late_deduction_per', 3))
        config.tax_percentage = float(request.POST.get('tax_percentage', 0))
        config.provident_fund_pct = float(request.POST.get('provident_fund_pct', 0))
        config.security_pct = float(request.POST.get('security_pct', 0))
        config.van_child_pct = float(request.POST.get('van_child_pct', 0))
        config.new_employee_security_pct = float(request.POST.get('new_employee_security_pct', 50))
        config.housing_allowance_pct = float(request.POST.get('housing_allowance_pct', 0))
        config.medical_allowance_pct = float(request.POST.get('medical_allowance_pct', 0))
        config.transport_allowance_pct = float(request.POST.get('transport_allowance_pct', 0))
        config.fuel_allowance_pct = float(request.POST.get('fuel_allowance_pct', 0))
        config.bonus_per_day = float(request.POST.get('bonus_per_day', 0))
        config.bonus_percentage = float(request.POST.get('bonus_percentage', 0))
        config.save()

        # Create or update MonthlySalary records for ALL employees
        created_count = 0
        updated_count = 0
        all_employees = TeacherProfile.objects.filter(is_employee_separated=False)
        school = get_user_school(request.user)
        if school and not request.user.is_superuser:
            all_employees = all_employees.filter(school=school)
        for emp in all_employees:
            try:
                emp_salary, _ = EmployeeSalary.objects.get_or_create(
                    employee=emp, defaults={'basic_salary': emp.salary}
                )
                if emp.salary > 0 and emp_salary.basic_salary != emp.salary:
                    emp_salary.basic_salary = emp.salary

                # Reset custom overrides to use global config
                emp_salary.use_custom_config = False
                emp_salary.custom_housing_pct = config.housing_allowance_pct
                emp_salary.custom_medical_pct = config.medical_allowance_pct
                emp_salary.custom_transport_pct = config.transport_allowance_pct
                emp_salary.custom_fuel_pct = config.fuel_allowance_pct
                emp_salary.custom_tax_pct = config.tax_percentage
                emp_salary.custom_pf_pct = config.provident_fund_pct
                emp_salary.custom_security_pct = config.security_pct
                emp_salary.custom_van_child_pct = config.van_child_pct
                emp_salary.custom_bonus_per_day = config.bonus_per_day
                emp_salary.custom_bonus_pct = config.bonus_percentage
                emp_salary.save()

                ms, created = MonthlySalary.objects.get_or_create(
                    employee=emp, month=month, year=year,
                    defaults={
                        'salary_config': config,
                        'total_working_days': config.default_working_days,
                        'basic_salary': emp.salary,
                    }
                )
                if created:
                    created_count += 1
                else:
                    ms.salary_config = config
                    ms.total_working_days = config.default_working_days
                    ms.save()
                    updated_count += 1
            except Exception as e:
                import traceback
                traceback.print_exc()

        messages.success(request, f'Salary configuration saved. {created_count} new salary sheets created, {updated_count} updated.')
        return redirect(f'/admin-console/?section=salary-config&month={month}&year={year}')

    return redirect(f'/admin-console/?section=salary-config&month={month}&year={year}')


@login_required(login_url='admin_login')
def save_employee_overrides(request):
    """Save per-employee salary config overrides for a specific month."""
    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    if request.method != 'POST':
        return redirect('admin_console')

    month = int(request.POST.get('month', timezone.now().month))
    year = int(request.POST.get('year', timezone.now().year))
    config = SalaryConfig.objects.filter(month=month, year=year).first()
    if not config:
        config = SalaryConfig.objects.create(month=month, year=year)

    school = get_user_school(request.user)
    employees = TeacherProfile.objects.filter(is_employee_separated=False)
    if school and not request.user.is_superuser:
        employees = employees.filter(school=school)

    def _run_migrate():
        from django.core.management import call_command
        call_command('migrate', verbosity=0)

    def _save_one(emp):
        emp_id = str(emp.id)
        if f'custom_housing_{emp_id}' not in request.POST:
            return False

        ms, _ = MonthlySalary.objects.get_or_create(
            employee=emp, month=month, year=year,
            defaults={
                'salary_config': config,
                'total_working_days': config.default_working_days,
                'basic_salary': emp.salary,
            }
        )
        ms.cfg_housing_pct = float(request.POST.get(f'custom_housing_{emp_id}', 0))
        ms.cfg_medical_pct = float(request.POST.get(f'custom_medical_{emp_id}', 0))
        ms.cfg_transport_pct = float(request.POST.get(f'custom_transport_{emp_id}', 0))
        ms.cfg_fuel_pct = float(request.POST.get(f'custom_fuel_{emp_id}', 0))
        ms.cfg_tax_pct = float(request.POST.get(f'custom_tax_{emp_id}', 0))
        ms.cfg_pf_pct = float(request.POST.get(f'custom_pf_{emp_id}', 0))
        ms.cfg_security_pct = float(request.POST.get(f'custom_security_{emp_id}', 0))
        ms.cfg_van_child_pct = float(request.POST.get(f'custom_van_child_{emp_id}', 0))
        ms.cfg_bonus_per_day = float(request.POST.get(f'custom_bonus_per_day_{emp_id}', 0))
        ms.cfg_bonus_pct = float(request.POST.get(f'custom_bonus_pct_{emp_id}', 0))
        ms.transaction_type = request.POST.get(f'custom_transaction_type_{emp_id}', 'bank_islami')
        ms.has_custom_config = True
        ms.save()
        return True

    saved_count = 0
    migrated = False
    error_msgs = []
    for emp in employees:
        try:
            if _save_one(emp):
                saved_count += 1
        except Exception as e:
            err_str = str(e).lower()
            if not migrated and ('column' in err_str or 'does not exist' in err_str or 'operational' in err_str):
                try:
                    _run_migrate()
                    migrated = True
                    if _save_one(emp):
                        saved_count += 1
                except Exception as e2:
                    import traceback
                    traceback.print_exc()
                    error_msgs.append(f'{emp.full_name}: {e2}')
            else:
                import traceback
                traceback.print_exc()
                error_msgs.append(f'{emp.full_name}: {e}')

    if error_msgs:
        messages.warning(request, f'Errors for {len(error_msgs)} employees: {"; ".join(error_msgs[:3])}')
    messages.success(request, f'Salary config saved for {saved_count} employees.')
    return redirect(f'/admin-console/?section=salary-config&month={month}&year={year}')


# ─────────────────────────────────────────────
# EMPLOYEE SALARY STRUCTURE (Add/Edit/Delete)
# ─────────────────────────────────────────────

@login_required(login_url='admin_login')
def add_employee_salary(request, employee_id):
    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    employee = get_object_or_404(TeacherProfile, pk=employee_id)
    if request.method == 'POST':
        form = EmployeeSalaryForm(request.POST)
        if form.is_valid():
            sal = form.save(commit=False)
            sal.employee = employee
            sal.save()
            messages.success(request, f'Salary structure saved for {employee.full_name}.')
            return redirect('admin_console')
    else:
        form = EmployeeSalaryForm(initial={'employee': employee})
    return render(request, 'hr/employee_salary_form.html', {
        'form': form, 'employee': employee, 'action': 'Add'
    })


@login_required(login_url='admin_login')
def edit_employee_salary(request, employee_id):
    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    employee = get_object_or_404(TeacherProfile, pk=employee_id)
    sal, created = EmployeeSalary.objects.get_or_create(employee=employee)

    if request.method == 'POST':
        form = EmployeeSalaryForm(request.POST, instance=sal)
        if form.is_valid():
            form.save()
            messages.success(request, f'Salary structure updated for {employee.full_name}.')
            return redirect('admin_console')
    else:
        form = EmployeeSalaryForm(instance=sal)
    return render(request, 'hr/employee_salary_form.html', {
        'form': form, 'employee': employee, 'action': 'Edit'
    })


@login_required(login_url='admin_login')
def delete_employee_salary(request, employee_id):
    """Delete salary structure for an employee."""
    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    employee = get_object_or_404(TeacherProfile, pk=employee_id)
    sal = EmployeeSalary.objects.filter(employee=employee).first()
    if sal:
        sal.delete()
        messages.success(request, f'Salary structure deleted for {employee.full_name}.')
    return redirect('admin_console')


# ─────────────────────────────────────────────
# GENERATE MONTHLY SALARY — Bulk
# ─────────────────────────────────────────────

@login_required(login_url='admin_login')
def generate_monthly_salary(request):
    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    school = get_user_school(request.user)
    employees = TeacherProfile.objects.filter(is_employee_separated=False)
    if school and not request.user.is_superuser:
        employees = employees.filter(school=school)

    month = int(request.POST.get('month', timezone.now().month)) if request.method == 'POST' else int(request.GET.get('month', timezone.now().month))
    year = int(request.POST.get('year', timezone.now().year)) if request.method == 'POST' else int(request.GET.get('year', timezone.now().year))
    config = SalaryConfig.objects.filter(month=month, year=year).first()

    if request.method == 'POST':
        if not config:
            messages.error(request, f'No salary configuration found for {calendar.month_name[month]} {year}. Please save salary configuration first.')
            return redirect('admin_console')
        month = int(request.POST.get('month', timezone.now().month))
        year = int(request.POST.get('year', timezone.now().year))
        total_working_days = int(request.POST.get('total_working_days', config.default_working_days))
        bonus_per_day = request.POST.get('bonus_per_day', '0')
        bonus_per_day = float(bonus_per_day) if bonus_per_day else 0

        selected_ids = request.POST.getlist('selected_employees')

        created_count = 0
        updated_count = 0
        skipped_count = 0

        for emp in employees:
            if selected_ids and str(emp.id) not in selected_ids:
                continue

            emp_salary, _ = EmployeeSalary.objects.get_or_create(
                employee=emp,
                defaults={'basic_salary': emp.salary}
            )
            if emp.salary > 0 and emp_salary.basic_salary != emp.salary:
                emp_salary.basic_salary = emp.salary
                emp_salary.save()

            monthly, created = MonthlySalary.objects.get_or_create(
                employee=emp, month=month, year=year,
                defaults={
                    'salary_config': config,
                    'total_working_days': total_working_days,
                    'days_absent': 0,
                    'unpaid_leaves': 0,
                    'paid_leaves': 0,
                    'late_coming_days': 0,
                    'basic_salary': emp.salary,
                    'increment': 0,
                    'bonus_per_day': bonus_per_day,
                }
            )
            if created:
                created_count += 1
            else:
                monthly.salary_config = config
                monthly.total_working_days = total_working_days
                monthly.save()
                updated_count += 1

        month_name = calendar.month_name[month]
        messages.success(request, f'Salary generated for {month_name} {year}: {created_count} new, {updated_count} updated.')
        return redirect('admin_console')

    return redirect('admin_console')


# ─────────────────────────────────────────────
# MONTHLY SALARY LIST (with search, filter, bulk actions)
# ─────────────────────────────────────────────

@login_required(login_url='admin_login')
def monthly_salary_list(request):
    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    month = request.GET.get('month', timezone.now().month)
    year = request.GET.get('year', timezone.now().year)
    try:
        month = int(month)
        year = int(year)
    except (ValueError, TypeError):
        month = timezone.now().month
        year = timezone.now().year

    salaries = MonthlySalary.objects.filter(month=month, year=year).select_related('employee')

    # Search
    search = request.GET.get('search', '').strip()
    if search:
        salaries = salaries.filter(
            Q(employee__full_name__icontains=search) |
            Q(employee__employee_id__icontains=search)
        )

    # Filter by pay status
    status_filter = request.GET.get('status', '')
    if status_filter:
        salaries = salaries.filter(pay_status=status_filter)

    month_name = calendar.month_name[month]

    # Bulk mark paid
    if request.method == 'POST':
        action = request.POST.get('action', '')
        selected_ids = request.POST.getlist('selected_salaries')
        if action == 'bulk_mark_paid' and selected_ids:
            payment_date = request.POST.get('payment_date', str(timezone.now().date()))
            from datetime import date as dt_date
            try:
                pay_date = dt_date.fromisoformat(payment_date)
            except ValueError:
                pay_date = timezone.now().date()
            count = MonthlySalary.objects.filter(pk__in=selected_ids).update(
                pay_status='paid', payment_date=pay_date
            )
            messages.success(request, f'{count} salary records marked as paid.')
            return redirect('admin_console')
        elif action == 'bulk_delete' and selected_ids:
            count = MonthlySalary.objects.filter(pk__in=selected_ids).delete()[0]
            messages.success(request, f'{count} salary records deleted.')
            return redirect('admin_console')

    # Totals
    total_gross = sum(s.gross_salary for s in salaries)
    total_deductions = sum(s.total_deductions for s in salaries)
    total_net = sum(s.net_salary for s in salaries)
    total_basic = sum(s.basic_salary for s in salaries)

    return redirect('admin_console')


# ─────────────────────────────────────────────
# EDIT MONTHLY SALARY
# ─────────────────────────────────────────────

@login_required(login_url='admin_login')
def edit_monthly_salary(request, pk):
    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    salary = get_object_or_404(MonthlySalary, pk=pk)

    if request.method == 'POST':
        salary.total_working_days = int(request.POST.get('total_working_days', salary.total_working_days))
        salary.days_absent = int(request.POST.get('days_absent', salary.days_absent))
        salary.paid_leaves = int(request.POST.get('paid_leaves', salary.paid_leaves))
        salary.unpaid_leaves = int(request.POST.get('unpaid_leaves', salary.unpaid_leaves))
        salary.late_coming_days = int(request.POST.get('late_coming_days', salary.late_coming_days))
        salary.increment = float(request.POST.get('increment', salary.increment))
        salary.housing_allowance = float(request.POST.get('housing_allowance', salary.housing_allowance))
        salary.medical_allowance = float(request.POST.get('medical_allowance', salary.medical_allowance))
        salary.transport_allowance = float(request.POST.get('transport_allowance', salary.transport_allowance))
        salary.fuel_allowance = float(request.POST.get('fuel_allowance', salary.fuel_allowance))
        salary.other_allowance = float(request.POST.get('other_allowance', salary.other_allowance))
        salary.advance_deduction = float(request.POST.get('advance_deduction', salary.advance_deduction))
        salary.provident_fund = float(request.POST.get('provident_fund', salary.provident_fund))
        salary.security_deduction = float(request.POST.get('security_deduction', salary.security_deduction))
        salary.van_child_deduction = float(request.POST.get('van_child_deduction', salary.van_child_deduction))
        salary.other_deduction = float(request.POST.get('other_deduction', salary.other_deduction))
        salary.overtime_hours = float(request.POST.get('overtime_hours', salary.overtime_hours))
        salary.overtime_rate = float(request.POST.get('overtime_rate', salary.overtime_rate))
        salary.remarks = request.POST.get('remarks', salary.remarks)
        salary.pay_status = request.POST.get('pay_status', salary.pay_status)
        payment_date = request.POST.get('payment_date', '')
        if payment_date:
            from datetime import date as dt_date
            try:
                salary.payment_date = dt_date.fromisoformat(payment_date)
            except ValueError:
                pass
        salary.save()
        messages.success(request, f'Salary updated for {salary.employee.full_name}.')
        return redirect('admin_console')

    return render(request, 'hr/edit_monthly_salary.html', {'salary': salary})


# ─────────────────────────────────────────────
# DELETE MONTHLY SALARY
# ─────────────────────────────────────────────

@login_required(login_url='admin_login')
def delete_monthly_salary(request, pk):
    """Delete a monthly salary record."""
    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    salary = get_object_or_404(MonthlySalary, pk=pk)
    month = salary.month
    year = salary.year
    name = salary.employee.full_name
    salary.delete()
    messages.success(request, f'Salary record deleted for {name}.')
    return redirect('admin_console')


# ─────────────────────────────────────────────
# CSV EXPORT
# ─────────────────────────────────────────────

@login_required(login_url='admin_login')
def export_salary_csv(request):
    """Export salary sheet as CSV for bank transfer."""
    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    month = int(request.GET.get('month', timezone.now().month))
    year = int(request.GET.get('year', timezone.now().year))

    salaries = MonthlySalary.objects.filter(month=month, year=year).select_related('employee')
    month_name = calendar.month_name[month]

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="Salary_{month_name}_{year}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Employee Name', 'Employee ID', 'Designation', 'Bank Name', 'Account Number',
        'Basic Salary', 'Total Allowances', 'Gross Salary', 'Total Deductions',
        'Net Salary', 'Pay Status'
    ])
    for s in salaries:
        emp = s.employee
        emp_sal = EmployeeSalary.objects.filter(employee=emp).first()
        writer.writerow([
            emp.full_name, emp.employee_id or '', emp.designation or '',
            emp_sal.bank_name if emp_sal else '', emp_sal.bank_account if emp_sal else '',
            s.basic_salary, s.total_allowances, s.gross_salary,
            s.total_deductions, s.net_salary, s.get_pay_status_display()
        ])

    return response


# ─────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────

@login_required(login_url='admin_login')
def export_salary_excel(request):
    """Export salary sheet as Excel for the selected month/year."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    month = int(request.GET.get('month', timezone.now().month))
    year = int(request.GET.get('year', timezone.now().year))

    salaries = MonthlySalary.objects.filter(month=month, year=year).select_related('employee')
    month_name = calendar.month_name[month]

    wb = Workbook()
    ws = wb.active
    ws.title = f"Salary {month_name} {year}"

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    headers = [
        'Sr#', 'Employee Name', 'Employee ID', 'Designation',
        'Basic Salary', 'Housing Allowance', 'Medical Allowance', 'Transport Allowance',
        'Fuel Allowance', 'Bonus', 'Gross Salary',
        'Tax Deduction', 'Provident Fund', 'Security Deduction', 'Van/Child Deduction',
        'Leave Deduction', 'Late Deduction',
        'Total Deductions', 'Net Salary', 'Transaction Type',
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    data_font = Font(name='Calibri', size=10)
    amount_font = Font(name='Calibri', size=10, bold=True)
    green_font = Font(name='Calibri', size=10, bold=True, color='16A34A')
    red_font = Font(name='Calibri', size=10, bold=True, color='DC2626')
    amount_alignment = Alignment(horizontal='right')

    for row_num, s in enumerate(salaries, 2):
        emp = s.employee

        row_data = [
            row_num - 1,
            emp.full_name,
            emp.employee_id or '',
            emp.designation or '',
            float(s.basic_salary),
            float(s.housing_allowance),
            float(s.medical_allowance),
            float(s.transport_allowance),
            float(s.fuel_allowance),
            float(s.bonus_amount),
            float(s.gross_salary),
            float(s.tax_deduction),
            float(s.provident_fund),
            float(s.security_deduction),
            float(s.van_child_deduction),
            float(s.leave_deduction),
            float(s.late_coming_deduction),
            float(s.total_deductions),
            float(s.net_salary),
            s.get_transaction_type_display() if hasattr(s, 'get_transaction_type_display') else '',
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            if col_num >= 5 and col_num <= 19:
                cell.font = amount_font
                cell.alignment = amount_alignment
                cell.number_format = '#,##0'
            else:
                cell.font = data_font

        ws.cell(row=row_num, column=18).font = green_font
        ws.cell(row=row_num, column=17).font = red_font

    if salaries:
        total_row = len(salaries) + 2
        ws.cell(row=total_row, column=1, value='TOTAL').font = Font(name='Calibri', bold=True, size=11)
        ws.cell(row=total_row, column=1).border = thin_border

        total_fields = {
            5: 'basic_salary', 6: 'housing_allowance', 7: 'medical_allowance',
            8: 'transport_allowance', 9: 'fuel_allowance',
            10: 'bonus_amount', 11: 'gross_salary', 12: 'tax_deduction',
            13: 'provident_fund', 14: 'security_deduction', 15: 'van_child_deduction',
            16: 'leave_deduction', 17: 'late_coming_deduction',
            18: 'total_deductions', 19: 'net_salary',
        }
        for col_num, field in total_fields.items():
            total_val = sum(getattr(s, field) for s in salaries)
            cell = ws.cell(row=total_row, column=col_num, value=float(total_val))
            cell.font = Font(name='Calibri', bold=True, size=10)
            cell.alignment = amount_alignment
            cell.number_format = '#,##0'
            cell.border = thin_border

    col_widths = [5, 22, 14, 18, 14, 14, 14, 14, 12, 10, 14, 12, 12, 14, 14, 12, 12, 14, 14, 16]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Salary_{month_name}_{year}.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────
# EXCEL IMPORT
# ─────────────────────────────────────────────

@login_required(login_url='admin_login')
def import_salary_excel(request):
    """Import salary data from an uploaded Excel file for the selected month/year."""
    from openpyxl import load_workbook

    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    if request.method != 'POST':
        return redirect('admin_console')

    month = int(request.POST.get('month', timezone.now().month))
    year = int(request.POST.get('year', timezone.now().year))
    excel_file = request.FILES.get('excel_file')

    if not excel_file:
        messages.error(request, 'Please select an Excel file to import.')
        return redirect(f'/admin-console/?section=salary-sheet&month={month}&year={year}')

    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'Invalid file format. Please upload an .xlsx or .xls file.')
        return redirect(f'/admin-console/?section=salary-sheet&month={month}&year={year}')

    try:
        wb = load_workbook(excel_file)
        ws = wb.active

        header_row = [str(cell.value or '').strip().lower() for cell in ws[1]]

        col_map = {}
        for idx, h in enumerate(header_row):
            if 'employee name' in h or h == 'teacher':
                col_map['employee_name'] = idx
            elif 'employee id' in h:
                col_map['employee_id'] = idx
            elif 'basic' in h and 'salary' in h:
                col_map['basic_salary'] = idx
            elif 'housing' in h:
                col_map['housing_allowance'] = idx
            elif 'medical' in h:
                col_map['medical_allowance'] = idx
            elif 'transport' in h:
                col_map['transport_allowance'] = idx
            elif 'fuel' in h:
                col_map['fuel_allowance'] = idx
            elif 'other allowance' in h:
                col_map['other_allowance'] = idx
            elif 'bonus' in h:
                col_map['bonus_amount'] = idx
            elif 'gross' in h:
                col_map['gross_salary'] = idx
            elif 'tax' in h:
                col_map['tax_deduction'] = idx
            elif 'provident' in h or h == 'pf':
                col_map['provident_fund'] = idx
            elif 'security' in h and 'deduction' in h:
                col_map['security_deduction'] = idx
            elif 'van' in h or 'child' in h:
                col_map['van_child_deduction'] = idx
            elif 'leave ded' in h:
                col_map['leave_deduction'] = idx
            elif 'late ded' in h:
                col_map['late_coming_deduction'] = idx
            elif 'advance' in h:
                col_map['advance_deduction'] = idx
            elif 'other ded' in h:
                col_map['other_deduction'] = idx
            elif 'total ded' in h:
                col_map['total_deductions'] = idx
            elif 'net' in h:
                col_map['net_salary'] = idx
            elif 'pay status' in h:
                col_map['pay_status'] = idx
            elif 'transaction' in h:
                col_map['transaction_type'] = idx

        if 'employee_name' not in col_map and 'employee_id' not in col_map:
            messages.error(request, f'Could not find Employee Name or Employee ID column. Detected columns: {", ".join(header_row)}')
            return redirect(f'/admin-console/?section=salary-sheet&month={month}&year={year}')

        def get_val(row, key, default=0):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return default
            val = row[idx]
            if val is None:
                return default
            try:
                return float(val)
            except (ValueError, TypeError):
                return default

        def get_str(row, key, default=''):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return default
            val = row[idx]
            if val is None:
                return default
            if isinstance(val, float):
                val = str(int(val)) if val == int(val) else str(val)
            return str(val).strip()

        imported = 0
        updated = 0
        skipped = 0
        skipped_names = []
        errors = []

        for row_num in range(2, ws.max_row + 1):
            row = [cell.value for cell in ws[row_num]]
            name_val = get_str(row, 'employee_name')
            id_val = get_str(row, 'employee_id')

            if not name_val and not id_val:
                continue

            emp = None
            if id_val:
                emp = TeacherProfile.objects.filter(employee_id=id_val).first()
            if not emp and name_val:
                clean_name = name_val.strip()
                emp = TeacherProfile.objects.filter(full_name__iexact=clean_name).first()
            if not emp and name_val:
                clean_name = name_val.strip()
                emp = TeacherProfile.objects.filter(full_name__icontains=clean_name).first()
            if not emp and name_val:
                clean_name = name_val.strip()
                emp = TeacherProfile.objects.filter(phone__iexact=clean_name).first()
            if not emp and name_val:
                clean_name = name_val.strip()
                emp = TeacherProfile.objects.filter(cnic__iexact=clean_name).first()
            if not emp:
                skipped += 1
                if name_val and name_val not in skipped_names:
                    skipped_names.append(name_val)
                continue

            basic = get_val(row, 'basic_salary')
            if basic <= 0:
                basic = float(emp.salary) if emp.salary else 0

            pay_status = get_str(row, 'pay_status', 'unpaid').lower()
            if pay_status not in ('paid', 'unpaid', 'partial'):
                pay_status = 'unpaid'

            transaction_type = get_str(row, 'transaction_type', 'bank_islami')
            valid_transactions = ['bank_islami', 'ubl', 'cash', 'personal']
            if transaction_type.lower() not in valid_transactions:
                transaction_type = 'bank_islami'
            else:
                transaction_type = transaction_type.lower()

            ms, created = MonthlySalary.objects.get_or_create(
                employee=emp, month=month, year=year,
                defaults={
                    'total_working_days': 26,
                    'basic_salary': basic,
                }
            )

            ms.basic_salary = basic
            ms.housing_allowance = get_val(row, 'housing_allowance')
            ms.medical_allowance = get_val(row, 'medical_allowance')
            ms.transport_allowance = get_val(row, 'transport_allowance')
            ms.fuel_allowance = get_val(row, 'fuel_allowance')
            ms.other_allowance = get_val(row, 'other_allowance')
            ms.bonus_amount = get_val(row, 'bonus_amount')
            ms.tax_deduction = get_val(row, 'tax_deduction')
            ms.provident_fund = get_val(row, 'provident_fund')
            ms.security_deduction = get_val(row, 'security_deduction')
            ms.van_child_deduction = get_val(row, 'van_child_deduction')
            ms.leave_deduction = get_val(row, 'leave_deduction')
            ms.late_coming_deduction = get_val(row, 'late_coming_deduction')
            ms.advance_deduction = get_val(row, 'advance_deduction')
            ms.other_deduction = get_val(row, 'other_deduction')
            ms.pay_status = pay_status
            ms.transaction_type = transaction_type
            ms.save()

            if created:
                imported += 1
            else:
                updated += 1

        month_name = calendar.month_name[month]
        skipped_preview = ', '.join(skipped_names[:5])
        if len(skipped_names) > 5:
            skipped_preview += f'... and {len(skipped_names) - 5} more'
        msg = f'Import complete for {month_name} {year}: {imported} new, {updated} updated, {skipped} skipped.'
        if skipped_names:
            msg += f' Not found: {skipped_preview}'
        messages.success(request, msg)
    except Exception as e:
        import traceback
        traceback.print_exc()
        messages.error(request, f'Error importing Excel file: {str(e)}')

    return redirect(f'/admin-console/?section=salary-sheet&month={month}&year={year}')


# ─────────────────────────────────────────────
# SALARY SLIP (with dynamic school name)
# ─────────────────────────────────────────────

@login_required(login_url='admin_login')
def salary_slip(request, pk):
    salary = get_object_or_404(MonthlySalary, pk=pk)
    employee = salary.employee
    emp_salary = EmployeeSalary.objects.filter(employee=employee).first()
    month_name = calendar.month_name[salary.month]
    school = getattr(employee, 'school', None)

    context = {
        'salary': salary,
        'employee': employee,
        'emp_salary': emp_salary,
        'month_name': month_name,
        'school': school,
    }
    return render(request, 'hr/salary_slip.html', context)


@login_required(login_url='admin_login')
def salary_slip_pdf(request, pk):
    """Returns salary slip as a downloadable file."""
    salary = get_object_or_404(MonthlySalary, pk=pk)
    employee = salary.employee
    emp_salary = EmployeeSalary.objects.filter(employee=employee).first()
    month_name = calendar.month_name[salary.month]
    school = getattr(employee, 'school', None)

    context = {
        'salary': salary,
        'employee': employee,
        'emp_salary': emp_salary,
        'month_name': month_name,
        'school': school,
        'print_mode': True,
    }
    html_string = render(request, 'hr/salary_slip_print.html', context).content.decode('utf-8')

    filename = f"{employee.full_name}_Salary_{month_name}_{salary.year}.html"
    response = HttpResponse(html_string, content_type='text/html')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required(login_url='admin_login')
def salary_slip_all(request):
    """Print all salary slips for a given month/year."""
    month = int(request.GET.get('month', timezone.now().month))
    year = int(request.GET.get('year', timezone.now().year))

    salaries = MonthlySalary.objects.filter(month=month, year=year).select_related('employee')
    month_name = calendar.month_name[month]

    # Get school from first employee
    school = None
    if salaries:
        school = getattr(salaries[0].employee, 'school', None)

    return redirect('admin_console')


# ─────────────────────────────────────────────
# MONTHLY ATTENDANCE SUMMARY (replaces daily attendance)
# ─────────────────────────────────────────────

@login_required(login_url='admin_login')
def monthly_attendance_summary(request):
    """HR enters monthly attendance summary: absents, paid leaves, unpaid leaves, late days."""
    role = getattr(request.user, 'role', None)
    if not (request.user.is_superuser or role in ('admin', 'admin_manager', 'principal')):
        return HttpResponse("Unauthorized", status=403)

    school = get_user_school(request.user)
    employees = TeacherProfile.objects.filter(is_employee_separated=False)
    if school and not request.user.is_superuser:
        employees = employees.filter(school=school)

    now = timezone.now()
    if request.method == 'POST':
        month = int(request.POST.get('month', now.month))
        year = int(request.POST.get('year', now.year))
    else:
        month = int(request.GET.get('month', now.month))
        year = int(request.GET.get('year', now.year))
    month_name = calendar.month_name[month]
    days_in_month = calendar.monthrange(year, month)[1]

    config = SalaryConfig.objects.filter(month=month, year=year).first()
    total_working_days = config.default_working_days if config else 26

    # Get existing monthly salary records for this month
    existing = {}
    for ms in MonthlySalary.objects.filter(month=month, year=year, employee__in=employees):
        existing[ms.employee_id] = ms

    if request.method == 'POST':
        created_count = 0
        updated_count = 0
        error_count = 0

        for emp in employees:
            try:
                days_absent = int(request.POST.get(f'absent_{emp.id}', 0))
                paid_leaves = int(request.POST.get(f'paid_leaves_{emp.id}', 0))
                unpaid_leaves = int(request.POST.get(f'unpaid_leaves_{emp.id}', 0))
                late_days = int(request.POST.get(f'late_{emp.id}', 0))
                remarks = request.POST.get(f'remarks_{emp.id}', '')

                days_present = max(0, days_in_month - days_absent)

                existing_ms = MonthlySalary.objects.filter(employee=emp, month=month, year=year).first()
                if existing_ms:
                    existing_ms.salary_config = config
                    existing_ms.total_working_days = total_working_days
                    existing_ms.days_absent = days_absent
                    existing_ms.paid_leaves = paid_leaves
                    existing_ms.unpaid_leaves = unpaid_leaves
                    existing_ms.late_coming_days = late_days
                    existing_ms.days_present = days_present
                    existing_ms.remarks = remarks
                    existing_ms.save()
                    updated_count += 1
                else:
                    ms = MonthlySalary(
                        employee=emp,
                        salary_config=config,
                        month=month,
                        year=year,
                        total_working_days=total_working_days,
                        days_absent=days_absent,
                        paid_leaves=paid_leaves,
                        unpaid_leaves=unpaid_leaves,
                        late_coming_days=late_days,
                        days_present=days_present,
                        remarks=remarks,
                    )
                    ms.save()
                    created_count += 1
            except Exception as e:
                error_count += 1
                import traceback
                traceback.print_exc()

        if error_count:
            messages.warning(request, f'Attendance saved with {error_count} errors. {created_count} new, {updated_count} updated.')
        else:
            messages.success(request, f'Attendance saved for {month_name} {year}: {created_count} new, {updated_count} updated.')
        return redirect('admin_console')

    context = {
        'employees': employees,
        'month': month,
        'year': year,
        'month_name': month_name,
        'total_working_days': total_working_days,
        'existing': existing,
        'section': 'attendance',
    }
    return render(request, 'hr/monthly_attendance_summary.html', context)
